# Файл: /home/detker/Документы/repository/doc-forge/src/converters/xlsx_engine.py
"""
Движок конвертации таблиц Excel (.xlsx, .xls) в Markdown с сохранением структуры листов.
"""

from pathlib import Path
from typing import Dict, Any, List
import openpyxl

from src.converters.base import BaseConverter, ConversionResult


class XlsxConverter(BaseConverter):
    """
    Конвертер Excel (.xlsx) -> Markdown.
    """

    def can_handle(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx"]

    def convert(self, file_path: Path, extract_profile: bool = False) -> ConversionResult:
        if not file_path.exists():
            raise FileNotFoundError(f"Файл таблицы '{file_path}' не найден!")

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        md_sections: List[str] = []
        total_tables = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            # Фильтруем полностью пустые строки
            non_empty_rows = []
            for r in rows:
                if any(cell is not None and str(cell).strip() != "" for cell in r):
                    clean_row = [str(cell if cell is not None else "").strip() for cell in r]
                    non_empty_rows.append(clean_row)

            if not non_empty_rows:
                continue

            total_tables += 1
            md_sections.append(f"## Лист: {sheet_name}\n")

            # Выравниваем длину строк
            max_cols = max(len(r) for r in non_empty_rows)
            # Обрезаем завершающие пустые колонки
            while max_cols > 0 and all(r[max_cols - 1] == "" for r in non_empty_rows if len(r) >= max_cols):
                max_cols -= 1

            if max_cols == 0:
                continue

            headers = [r if r != "" else f"Колонка {i+1}" for i, r in enumerate(non_empty_rows[0][:max_cols])]
            md_sections.append("| " + " | ".join(headers) + " |")
            md_sections.append("| " + " | ".join([":---"] * max_cols) + " |")

            for r in non_empty_rows[1:]:
                padded_r = r[:max_cols] + [""] * (max_cols - len(r[:max_cols]))
                # Экранируем переносы строк внутри ячеек
                escaped_r = [c.replace("\n", " ").replace("|", "\\|") for c in padded_r]
                md_sections.append("| " + " | ".join(escaped_r) + " |")

            md_sections.append("\n")

        full_md = f"# Книга Excel: {file_path.name}\n\n" + "\n".join(md_sections)

        layout_report = (
            f"📊 ТАБЛИЦА EXCEL ({file_path.name}):\n"
            f"   • Количество листов с данными: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})\n"
            f"   • Таблиц сформировано: {total_tables}"
        )

        return ConversionResult(
            markdown=full_md,
            profile=None,
            metadata={
                "source_file": str(file_path),
                "sheet_names": wb.sheetnames,
                "tables_count": total_tables,
                "format": "XLSX",
            },
            layout_report=layout_report,
            tables_count=total_tables,
            pages_count=len(wb.sheetnames),
        )
